# -*- coding: utf-8 -*-
"""
Created on Wed May 24 15:11:15 2023

@author: CH226683
"""

import openpyxl
import os
from openpyxl import load_workbook
import csv

# 'from StyleFrame import StyleFrame' in older versions (< 3.0)

#xlsx_path = 'G:\\My Drive\\BCH\\Rotations\\Informatics\\CT wiki\\Siemens_Force\\2020-06 vs 2022-02-09\\'
#xlsx_path = 'G:\\My Drive\\BCH\\Rotations\\Informatics\\CT wiki\\Siemens_Force\\2020-06 vs 2023-08-10\\'
xlsx_path = 'G:\\My Drive\\BCH\\Rotations\\Informatics\\CT wiki\\Siemens_Force\\2022-02-09 vs 2023-08-10\\'
#xlsx_path = 'G:\\My Drive\\BCH\\Rotations\\Informatics\\CT wiki\\Siemens_Force\\2022-02-09 vs 2022-09-28\\'
#xlsx_path = 'G:\\My Drive\\BCH\\Rotations\\Informatics\\CT wiki\\Siemens_Force\\2022-09-28 vs 2023-04-25\\'
# xlsx_path = 'G:\\My Drive\\BCH\\Rotations\\Informatics\\CT wiki\\Siemens_Force\\2023-04-25 vs 2023-08-10\\'

onlyfiles = [os.path.join(dirpath,f) for (dirpath, dirnames, filenames) in os.walk(xlsx_path) for f in filenames]

# fields = []

wb_new = openpyxl.Workbook()
ws_write = wb_new.create_sheet()
# ws_write.append(['Protocol','Series/Recon','Field','06/2020','02/2022'])
#ws_write.append(['Protocol','Series/Recon','Field','06/2020','08/2023'])
ws_write.append(['Protocol','Series/Recon','Field','02/2022','08/2023'])
#ws_write.append(['Protocol','Series/Recon','Field','02/2022','09/2022'])
#ws_write.append(['Protocol','Series/Recon','Field','09/2022','04/2023'])
# ws_write.append(['Protocol','Series/Recon','Field','04/2023','08/2023'])
header_font = openpyxl.styles.Font(color='00FF0000', bold=True)
for cell in ws_write["1:1"]:
    cell.font = header_font

a_count = 0
c_count = 0

l_count = 0
ms_count = 0
r_count = -1 # due to header column being labeled "Series/Recon"
mon_count = 0
pmon_count = 0

for x in onlyfiles:
    wb = load_workbook(filename = x)
    ws = wb.get_sheet_by_name('Sheet1')
    
#    for row in ws.iter_rows():
    # for row in ws.iter_rows(min_row=2,min_col=2):
    #     for cell in row:
    #         #if cell.font.color != '#9c0006':
    #         # if cell.font.color == '#9c0006':
    #         # if cell.font.bold == True:
    #             fields.append(cell.value)
    #            print(cell.value)
    for row in ws.iter_rows(min_row=1,min_col=1):
        if type(row[1]).__name__ == 'MergedCell':
            hold_merge = row[0].value
        if row[1].value != row[2].value:
            #if cell.font.color != '#9c0006':
            # if cell.font.color == '#9c0006':
            # if cell.font.bold == True:
                # fields.append([ws.cell(1,2).value,hold_merge,row[0].value,row[1].value,row[2].value])
                ws_write.append([ws.cell(1,2).value,hold_merge,row[0].value,row[1].value,row[2].value])

mylist = []
for cell in ws_write['A']:
    mylist.append(cell.value)
    
    if '.Adult' in cell.value:
        a_count += 1
    elif '.Child' in cell.value:
        c_count += 1
    # elif 'Recon' in cell.value:
    #     r_count += 1
    # elif cell.value == 'Monitoring Scan':
    #     mon_count += 1
    # elif cell.value == 'PreMonitoring Scan':
    #     pmon_count += 1

mergecount=0
startcell=1
for row in range(1, len(mylist)):
    #print(row, mylist[row-1], mylist[row])
    if mylist[row-1] == mylist[row]:
        mergecount += 1
    else:
        #print(row, mylist[row-1], mylist[row], startcell, mergecount)
        if mergecount > 0:
            ws_write.merge_cells(start_row=startcell, start_column=1, end_row=startcell+mergecount, end_column=1)
        mergecount = 0
        startcell = row+1
if mergecount > 0:
    ws_write.merge_cells(start_row=startcell, start_column=1, end_row=startcell+mergecount, end_column=1)

mylist = []
for cell in ws_write['B']:
    mylist.append(cell.value)
    
    if 'Localizer' in cell.value:
        l_count += 1
    elif 'Main Scan' in cell.value:
        ms_count += 1
    elif 'Recon' in cell.value:
        r_count += 1
    elif cell.value == 'Monitoring Scan':
        mon_count += 1
    elif cell.value == 'PreMonitoring Scan':
        pmon_count += 1
    
mergecount=0
startcell=1
for row in range(1, len(mylist)):
    #print(row, mylist[row-1], mylist[row])
    if mylist[row-1] == mylist[row]:
        mergecount += 1
    else:
        #print(row, mylist[row-1], mylist[row], startcell, mergecount)
        if mergecount > 0:
            ws_write.merge_cells(start_row=startcell, start_column=2, end_row=startcell+mergecount, end_column=2)
        mergecount = 0
        startcell = row+1
if mergecount > 0:
    ws_write.merge_cells(start_row=startcell, start_column=2, end_row=startcell+mergecount, end_column=2)

for row in range(1,ws_write.max_row+1):
    for col in range(1,ws_write.max_column+1):
        cell=ws_write.cell(row, col)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

for column_cells in ws_write.columns:
    new_column_length = max(len(str(cell.value)) for cell in column_cells)
    new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
    if new_column_length > 0:
        ws_write.column_dimensions[new_column_letter].width = new_column_length*1.23

del wb_new['Sheet']
#wb_new.save(filename='G:\\My Drive\\BCH\\Manuscripts\\CT Workflow\\ChangedParams_062020_022022.xlsx')
#wb_new.save(filename='G:\\My Drive\\BCH\\Manuscripts\\CT Workflow\\ChangedParams_062020_082023.xlsx')
wb_new.save(filename='G:\\My Drive\\BCH\\Manuscripts\\CT Workflow\\ChangedParams_022022_082023.xlsx')
#wb_new.save(filename='G:\\My Drive\\BCH\\Manuscripts\\CT Workflow\\ChangedParams_022022_092022.xlsx')
#wb_new.save(filename='G:\\My Drive\\BCH\\Manuscripts\\CT Workflow\\ChangedParams_092022_042023.xlsx')
# wb_new.save(filename='G:\\My Drive\\BCH\\Manuscripts\\CT Workflow\\ChangedParams_042023_082023.xlsx')

# with open('G:\\My Drive\\BCH\\Manuscripts\\CT Workflow\\ChangedParams.csv', 'w') as f:
      
#     # using csv.writer method from CSV package
#     write = csv.writer(f)
    
#     write.writerow(['Protocol','Series/Recon','Field','2022','2023'])
#     write.writerows(fields)

#     # for name in sf.Names:
#     #     if name.style.bold:
#     #         fields.append(name)

